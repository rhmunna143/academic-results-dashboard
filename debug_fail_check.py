import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Academic_Results_Dashboard.xlsx')
ws = wb['Data Source']

# Check row 3 (Fatima Khan)
row = 3
print(f"Analyzing Row {row}:")
print("=" * 60)

# Get all values
quran = ws[f'C{row}'].value
hadith = ws[f'D{row}'].value
arabic_i = ws[f'E{row}'].value
arabic_ii = ws[f'F{row}'].value
aqaid = ws[f'G{row}'].value
english_i = ws[f'H{row}'].value
english_ii = ws[f'I{row}'].value
bangla_i_mcq = ws[f'J{row}'].value
bangla_i_written = ws[f'K{row}'].value
bangla_ii_mcq = ws[f'L{row}'].value
bangla_ii_written = ws[f'M{row}'].value
math_mcq = ws[f'N{row}'].value
math_written = ws[f'O{row}'].value
history_mcq = ws[f'P{row}'].value
history_written = ws[f'Q{row}'].value
ict = ws[f'R{row}'].value
mantiq = ws[f'S{row}'].value
career = ws[f'T{row}'].value
physical = ws[f'U{row}'].value

print(f"\nSubject Marks:")
print(f"Quran: {quran}, Hadith: {hadith} → Total: {quran + hadith} (need ≥66)")
print(f"Arabic I: {arabic_i}, Arabic II: {arabic_ii} → Total: {arabic_i + arabic_ii} (need ≥66)")
print(f"Aqaid: {aqaid} (need ≥33)")
print(f"English I: {english_i}, English II: {english_ii} → Total: {english_i + english_ii} (need ≥66)")
print(f"Bangla I MCQ: {bangla_i_mcq}, Written: {bangla_i_written}")
print(f"Bangla II MCQ: {bangla_ii_mcq}, Written: {bangla_ii_written}")
print(f"  Total MCQ: {bangla_i_mcq + bangla_ii_mcq}, Total Written: {bangla_i_written + bangla_ii_written}")
print(f"Math MCQ: {math_mcq}, Written: {math_written}")
print(f"History MCQ: {history_mcq}, Written: {history_written}")
print(f"ICT: {ict} (need ≥8.25)")
print(f"Mantiq: {mantiq}")
print(f"Career: {career} (need ≥33)")
print(f"Physical: {physical} (need ≥33)")

print(f"\nFail Check Analysis:")
print("-" * 60)

# Check each fail condition
fail_conditions = []

if (quran + hadith) < 66:
    fail_conditions.append(f"Quran+Hadith: {quran + hadith} < 66 ❌")
else:
    print(f"Quran+Hadith: {quran + hadith} ≥ 66 ✓")

if (arabic_i + arabic_ii) < 66:
    fail_conditions.append(f"Arabic: {arabic_i + arabic_ii} < 66 ❌")
else:
    print(f"Arabic: {arabic_i + arabic_ii} ≥ 66 ✓")

if aqaid < 33:
    fail_conditions.append(f"Aqaid: {aqaid} < 33 ❌")
else:
    print(f"Aqaid: {aqaid} ≥ 33 ✓")

if (english_i + english_ii) < 66:
    fail_conditions.append(f"English: {english_i + english_ii} < 66 ❌")
else:
    print(f"English: {english_i + english_ii} ≥ 66 ✓")

# Bangla check
individual_pass = (bangla_i_mcq >= 10 and bangla_i_written >= 23 and bangla_ii_mcq >= 10 and bangla_ii_written >= 23)
combined_pass = ((bangla_i_mcq + bangla_ii_mcq) >= 20 and (bangla_i_written + bangla_ii_written) >= 46)
bangla_pass = individual_pass or combined_pass
if not bangla_pass:
    fail_conditions.append(f"Bangla: Failed both conditions ❌")
    print(f"  Individual: I_MCQ={bangla_i_mcq}≥10={bangla_i_mcq>=10}, I_Written={bangla_i_written}≥23={bangla_i_written>=23}, II_MCQ={bangla_ii_mcq}≥10={bangla_ii_mcq>=10}, II_Written={bangla_ii_written}≥23={bangla_ii_written>=23}")
    print(f"  Combined: MCQ={bangla_i_mcq + bangla_ii_mcq}≥20={(bangla_i_mcq + bangla_ii_mcq)>=20}, Written={bangla_i_written + bangla_ii_written}≥46={(bangla_i_written + bangla_ii_written)>=46}")
else:
    print(f"Bangla: Pass ✓ (Individual={individual_pass}, Combined={combined_pass})")

# Math check
if math_mcq < 10 or math_written < 23:
    fail_conditions.append(f"Math: MCQ={math_mcq}<10 or Written={math_written}<23 ❌")
else:
    print(f"Math: MCQ={math_mcq}≥10 and Written={math_written}≥23 ✓")

# History check
if history_mcq < 10 or history_written < 23:
    fail_conditions.append(f"History: MCQ={history_mcq}<10 or Written={history_written}<23 ❌")
else:
    print(f"History: MCQ={history_mcq}≥10 and Written={history_written}≥23 ✓")

if ict < 8.25:
    fail_conditions.append(f"ICT: {ict} < 8.25 ❌")
else:
    print(f"ICT: {ict} ≥ 8.25 ✓")

if career < 33:
    fail_conditions.append(f"Career: {career} < 33 ❌")
else:
    print(f"Career: {career} ≥ 33 ✓")

if physical < 33:
    fail_conditions.append(f"Physical: {physical} < 33 ❌")
else:
    print(f"Physical: {physical} ≥ 33 ✓")

print(f"\n{'FAILED' if fail_conditions else 'PASSED'} CONDITIONS:")
if fail_conditions:
    for condition in fail_conditions:
        print(f"  • {condition}")
else:
    print("  All pass conditions met!")

wb.close()
