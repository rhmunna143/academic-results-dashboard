import openpyxl

wb = openpyxl.load_workbook('Academic_Results_Dashboard.xlsx')
ws = wb['Data Source']

# Check formula lengths
gpa_formula = ws['X3'].value
grade_formula = ws['Y3'].value

print(f"GPA formula length: {len(gpa_formula)} characters")
print(f"Grade formula length: {len(grade_formula)} characters")
print(f"\nExcel's formula limit is ~8,192 characters")
print(f"But complex nested formulas can cause issues even below that limit\n")

if len(gpa_formula) > 1000:
    print("⚠️ GPA formula is very long and likely causing corruption!")
    
wb.close()
