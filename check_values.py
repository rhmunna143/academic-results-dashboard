import openpyxl

# Load with data_only=True to get calculated values
wb = openpyxl.load_workbook('Academic_Results_Dashboard.xlsx', data_only=True)
ws = wb['Data Source']

print('Checking calculated GPA and Grade values:')
print('-' * 50)
for i in range(3, 12):  # Check first 10 students
    gpa = ws[f'X{i}'].value
    grade = ws[f'Y{i}'].value
    name = ws[f'B{i}'].value
    print(f'Row {i} ({name}): GPA={gpa}, Grade={grade}')

wb.close()
