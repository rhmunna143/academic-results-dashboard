#!/usr/bin/env python3
"""
Academic Results Dashboard Generator
Creates an Excel file with Data Source, Pivot, and Dashboard sheets
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

def calculate_grade_point(marks, full_marks=100):
    """Calculate grade point for individual subject based on Dakhil curriculum
    
    Args:
        marks: Obtained marks
        full_marks: Total marks for the subject (100 or 200)
    """
    # Calculate percentage
    percentage = (marks / full_marks) * 100
    
    if percentage >= 80: return 5.0
    elif percentage >= 70: return 4.0
    elif percentage >= 60: return 3.5
    elif percentage >= 50: return 3.0
    elif percentage >= 40: return 2.0
    elif percentage >= 33: return 1.0
    else: return 0.0

def calculate_gpa_dakhil(row):
    """Calculate GPA based on Bangladeshi Dakhil curriculum
    
    Compulsory subjects:
    - 6 separate subjects (100 marks each)
    - 3 combined subjects (200 marks each)
    
    Additional subject (Mantiq):
    - Only counts if GPA >= 2.0
    - Adds bonus to final GPA
    
    Continuous Assessment (Career & Physical Education):
    - Pass/Fail only, doesn't affect GPA calculation
    """
    # Compulsory subjects for Dakhil curriculum (8 subjects, some split into MCQ+Written)
    # Combined subjects: Quran+Hadith, Arabic I+II, English I+II (2 columns each)
    # Bangla: 4 columns (Bangla_I_MCQ, Bangla_I_Written, Bangla_II_MCQ, Bangla_II_Written)
    # Math & Islamic History: 2 columns each (MCQ, Written)
    compulsory_subjects = {
        'Quran_Hadith': ('Quran', 'Hadith', 200),  # (col1, col2, total_marks)
        'Arabic': ('Arabic_I', 'Arabic_II', 200),
        'Aqaid': ('Aqaid', None, 100),  # Single column subjects
        'English': ('English_I', 'English_II', 200),
        'Bangla': ('Bangla_I_MCQ', None, 200),  # Will be handled specially (4 columns)
        'Mathematics': ('Mathematics_MCQ', None, 100),  # Will be handled specially (2 columns)
        'Islamic_History': ('Islamic_History_MCQ', None, 100),  # Will be handled specially (2 columns)
        'ICT': ('ICT', None, 50)
    }
    
    # Check if any compulsory subject failed (below 33%)
    for subject_name, subject_info in compulsory_subjects.items():
        col1, col2, full_marks = subject_info
        # If combined subject, add both columns
        if col2:
            marks = row[col1] + row[col2]
        else:
            marks = row[col1]
        
        # Special cases for pass thresholds
        if subject_name == 'ICT':
            # ICT: pass based on 33% of 25 marks (8.25)
            min_passing = 25 * 0.33
        elif subject_name == 'Bangla':
            # Bangla special: 4 columns (Bangla_I_MCQ, Bangla_I_Written, Bangla_II_MCQ, Bangla_II_Written)
            # Pass if: (Each MCQ≥10 AND Each Written≥23) OR (Total MCQ≥20 AND Total Written≥46)
            b1_mcq = row['Bangla_I_MCQ']
            b1_written = row['Bangla_I_Written']
            b2_mcq = row['Bangla_II_MCQ']
            b2_written = row['Bangla_II_Written']
            total_mcq = b1_mcq + b2_mcq
            total_written = b1_written + b2_written
            # Individual condition: Each part meets minimums
            individual_pass = (b1_mcq >= 10 and b1_written >= 23 and b2_mcq >= 10 and b2_written >= 23)
            # Combined condition: Totals meet minimums
            combined_pass = (total_mcq >= 20 and total_written >= 46)
            # Fail if neither condition is met
            if not (individual_pass or combined_pass):
                return 0.0
            # Calculate marks for GPA (total of all 4 columns)
            marks = b1_mcq + b1_written + b2_mcq + b2_written
            # Skip the general check below for Bangla
            continue
        elif subject_name == 'Mathematics':
            # Mathematics special: MCQ≥10 AND Written≥23
            mcq = row['Mathematics_MCQ']
            written = row['Mathematics_Written']
            if not (mcq >= 10 and written >= 23):
                return 0.0
            # Calculate marks for GPA
            marks = mcq + written
            # Skip the general check below
            continue
        elif subject_name == 'Islamic_History':
            # Islamic History special: MCQ≥10 AND Written≥23
            mcq = row['Islamic_History_MCQ']
            written = row['Islamic_History_Written']
            if not (mcq >= 10 and written >= 23):
                return 0.0
            # Calculate marks for GPA
            marks = mcq + written
            # Skip the general check below
            continue
        else:
            min_passing = full_marks * 0.33
        
        if marks < min_passing:
            return 0.0
    
    # Check continuous assessment subjects (must pass)
    if row['Career_Education'] < 33 or row['Physical_Education'] < 33:
        return 0.0
    
    # Calculate grade points for all compulsory subjects
    grade_points = []
    for subject_name, subject_info in compulsory_subjects.items():
        col1, col2, full_marks = subject_info
        # Handle special cases with multiple columns
        if subject_name == 'Bangla':
            marks = row['Bangla_I_MCQ'] + row['Bangla_I_Written'] + row['Bangla_II_MCQ'] + row['Bangla_II_Written']
        elif subject_name == 'Mathematics':
            marks = row['Mathematics_MCQ'] + row['Mathematics_Written']
        elif subject_name == 'Islamic_History':
            marks = row['Islamic_History_MCQ'] + row['Islamic_History_Written']
        # If combined subject, add both columns
        elif col2:
            marks = row[col1] + row[col2]
        else:
            marks = row[col1]
        gp = calculate_grade_point(marks, full_marks)
        grade_points.append(gp)
    
    # Calculate base GPA (average of compulsory subjects)
    base_gpa = sum(grade_points) / len(grade_points)
    
    # Handle additional subject (Mantiq) - Dakhil rule
    # If Mantiq GPA >= 2.0, add bonus points
    mantiq_gp = calculate_grade_point(row['Mantiq'], 100)
    if mantiq_gp >= 2.0:
        # Add (Mantiq GP - 2) / number of compulsory subjects
        bonus = (mantiq_gp - 2.0) / len(compulsory_subjects)
        base_gpa = base_gpa + bonus
    
    # Final GPA cannot exceed 5.0
    final_gpa = min(5.0, round(base_gpa, 2))
    return final_gpa

def calculate_letter_grade(gpa):
    """Convert GPA to letter grade"""
    if gpa >= 5.0: return 'A+'
    elif gpa >= 4.0: return 'A'
    elif gpa >= 3.5: return 'A-'
    elif gpa >= 3.0: return 'B'
    elif gpa >= 2.0: return 'C'
    elif gpa >= 1.0: return 'D'
    else: return 'F'

def create_data_source():
    """Create sample student data for Dakhil curriculum"""
    import random
    random.seed(42)  # For reproducible results
    
    data = {
        'SL': list(range(1, 21)),
        'Name': [
            'Ahmed Rahman', 'Fatima Khan', 'Karim Hassan', 'Ayesha Begum',
            'Rahim Uddin', 'Nadia Islam', 'Jahir Ahmed', 'Sadia Sultana',
            'Tarik Hasan', 'Ruhi Akter', 'Mehedi Hassan', 'Lubna Khatun',
            'Fahim Ahmed', 'Samira Begum', 'Imran Khan', 'Rafia Islam',
            'Shakib Rahman', 'Nusrat Jahan', 'Arif Hossain', 'Tasnia Akter'
        ],
        # Compulsory Subjects (8 subjects - 4 combined = 12 columns)
        'Quran': [random.randint(60, 95) for _ in range(20)],  # 100 marks
        'Hadith': [random.randint(65, 95) for _ in range(20)],  # 100 marks
        'Arabic_I': [random.randint(60, 95) for _ in range(20)],  # 100 marks
        'Arabic_II': [random.randint(65, 95) for _ in range(20)],  # 100 marks
        'Aqaid': [random.randint(60, 95) for _ in range(20)],  # 100 marks
        'English_I': [random.randint(55, 95) for _ in range(20)],  # 100 marks
        'English_II': [random.randint(60, 90) for _ in range(20)],  # 100 marks
        'Bangla_I_MCQ': [random.randint(8, 28) for _ in range(20)],  # 30 marks (pass: 10 individual or 20 combined)
        'Bangla_I_Written': [random.randint(20, 65) for _ in range(20)],  # 70 marks (pass: 23 individual or 46 combined)
        'Bangla_II_MCQ': [random.randint(8, 28) for _ in range(20)],  # 30 marks (pass: 10 individual or 20 combined)
        'Bangla_II_Written': [random.randint(20, 65) for _ in range(20)],  # 70 marks (pass: 23 individual or 46 combined)
        'Mathematics_MCQ': [random.randint(8, 28) for _ in range(20)],  # 30 marks (pass: 10)
        'Mathematics_Written': [random.randint(20, 65) for _ in range(20)],  # 70 marks (pass: 23)
        'Islamic_History_MCQ': [random.randint(8, 28) for _ in range(20)],  # 30 marks (pass: 10)
        'Islamic_History_Written': [random.randint(20, 65) for _ in range(20)],  # 70 marks (pass: 23)
        'ICT': [random.randint(30, 48) for _ in range(20)],  # 50 marks (pass on 33% of 25 = 8.25)
        # Additional Subject
        'Mantiq': [random.randint(50, 90) for _ in range(20)],  # 100 marks
        # Continuous Assessment Subjects
        'Career_Education': [random.randint(70, 95) for _ in range(20)],  # 100 marks (Pass/Fail)
        'Physical_Education': [random.randint(75, 95) for _ in range(20)],  # 100 marks (Pass/Fail)
    }
    
    # Return DataFrame with only raw marks. Total/Average/GPA/Grade will be
    # inserted into the workbook as Excel formulas so the workbook stays
    # dynamic when users edit marks directly in Excel.
    df = pd.DataFrame(data)
    subjects = ['Quran', 'Hadith', 'Arabic_I', 'Arabic_II', 'Aqaid', 'English_I', 'English_II', 
                'Bangla_I_MCQ', 'Bangla_I_Written', 'Bangla_II_MCQ', 'Bangla_II_Written', 
                'Mathematics_MCQ', 'Mathematics_Written', 'Islamic_History_MCQ', 'Islamic_History_Written', 
                'ICT', 'Mantiq', 'Career_Education', 'Physical_Education']
    return df[(['SL', 'Name'] + subjects)]

def style_data_source_sheet(ws, df):
    """Apply styling to Data Source sheet"""
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply conditional formatting colors for marks
    # All subjects are now 100 marks each (15 columns: C-Q)
    # C=Quran, D=Hadith, E=Arabic I, F=Arabic II, G=Aqaid, H=English I, I=English II, 
    # J=Bangla I, K=Bangla II, L=Math, M=Islamic History, N=ICT, O=Mantiq, P=Career, Q=Physical
    subject_cols_100_marks = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']

    for row in range(2, len(df) + 2):
        # All subjects are 100 marks
        for col in subject_cols_100_marks:
            cell = ws[f'{col}{row}']
            value = cell.value
            if value is None:
                continue
            try:
                num = float(value)
            except Exception:
                continue
            # Thresholds for 100 marks
            if num >= 80:
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            elif num >= 60:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif num >= 40:
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FC9999", end_color="FC9999", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5   # SL
    ws.column_dimensions['B'].width = 18  # Name
    ws.column_dimensions['C'].width = 10  # Quran
    ws.column_dimensions['D'].width = 10  # Hadith
    ws.column_dimensions['E'].width = 10  # Arabic I
    ws.column_dimensions['F'].width = 10  # Arabic II
    ws.column_dimensions['G'].width = 10  # Aqaid
    ws.column_dimensions['H'].width = 10  # English I
    ws.column_dimensions['I'].width = 10  # English II
    ws.column_dimensions['J'].width = 11  # Bangla I MCQ
    ws.column_dimensions['K'].width = 13  # Bangla I Written
    ws.column_dimensions['L'].width = 11  # Bangla II MCQ
    ws.column_dimensions['M'].width = 13  # Bangla II Written
    ws.column_dimensions['N'].width = 10  # Math MCQ
    ws.column_dimensions['O'].width = 12  # Math Written
    ws.column_dimensions['P'].width = 11  # History MCQ
    ws.column_dimensions['Q'].width = 13  # History Written
    ws.column_dimensions['R'].width = 10  # ICT
    ws.column_dimensions['S'].width = 10  # Mantiq
    ws.column_dimensions['T'].width = 12  # Career Education
    ws.column_dimensions['U'].width = 12  # Physical Education
    ws.column_dimensions['V'].width = 10  # Total
    ws.column_dimensions['W'].width = 10  # Average
    ws.column_dimensions['X'].width = 14  # GPA (Compulsory)
    ws.column_dimensions['Y'].width = 12  # GPA (Final)
    ws.column_dimensions['Z'].width = 12  # Overall Grade
    
    # Hide helper columns (AA through AO)
    for col in ['AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO']:
        ws.column_dimensions[col].hidden = True

def create_subject_grades_sheet(wb, df):
    """Create Subject Grades sheet with individual subject grades"""
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    ws = wb.create_sheet('Subject Grades')
    
    # Title
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = 'SUBJECT-WISE GRADES'
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['SL', 'Name', 'Quran+Hadith', 'Arabic', 'Aqaid', 'English', 'Bangla', 
               'Math', 'History', 'ICT', 'Mantiq', 'Career', 'Physical', 'Overall GPA', 'Overall Grade']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data and formulas
    for row_idx in range(len(df)):
        row = row_idx + 3  # Start from row 3 (after header)
        data_row = row_idx + 2  # Corresponding row in Data Source
        
        # SL and Name from Data Source
        ws[f'A{row}'] = f"='Data Source'!A{data_row}"
        ws[f'B{row}'] = f"='Data Source'!B{data_row}"
        
        # Subject grades with formulas (using new column layout)
        # Use simpler formulas by converting GP to letter grade
        # GP: 5=A+, 4=A, 3.5=A-, 3=B, 2=C, 1=D, 0=F
        
        # Quran+Hadith - use helper column AE from Data Source
        ws[f'C{row}'] = f"=IF('Data Source'!AE{data_row}>=5,\"A+\",IF('Data Source'!AE{data_row}>=4,\"A\",IF('Data Source'!AE{data_row}>=3.5,\"A-\",IF('Data Source'!AE{data_row}>=3,\"B\",IF('Data Source'!AE{data_row}>=2,\"C\",IF('Data Source'!AE{data_row}>=1,\"D\",\"F\"))))))"
        
        # Arabic - use helper column AF
        ws[f'D{row}'] = f"=IF('Data Source'!AF{data_row}>=5,\"A+\",IF('Data Source'!AF{data_row}>=4,\"A\",IF('Data Source'!AF{data_row}>=3.5,\"A-\",IF('Data Source'!AF{data_row}>=3,\"B\",IF('Data Source'!AF{data_row}>=2,\"C\",IF('Data Source'!AF{data_row}>=1,\"D\",\"F\"))))))"
        
        # Aqaid - use helper column AG
        ws[f'E{row}'] = f"=IF('Data Source'!AG{data_row}>=5,\"A+\",IF('Data Source'!AG{data_row}>=4,\"A\",IF('Data Source'!AG{data_row}>=3.5,\"A-\",IF('Data Source'!AG{data_row}>=3,\"B\",IF('Data Source'!AG{data_row}>=2,\"C\",IF('Data Source'!AG{data_row}>=1,\"D\",\"F\"))))))"
        
        # English - use helper column AH
        ws[f'F{row}'] = f"=IF('Data Source'!AH{data_row}>=5,\"A+\",IF('Data Source'!AH{data_row}>=4,\"A\",IF('Data Source'!AH{data_row}>=3.5,\"A-\",IF('Data Source'!AH{data_row}>=3,\"B\",IF('Data Source'!AH{data_row}>=2,\"C\",IF('Data Source'!AH{data_row}>=1,\"D\",\"F\"))))))"
        
        # Bangla - use helper column AI
        ws[f'G{row}'] = f"=IF('Data Source'!AI{data_row}>=5,\"A+\",IF('Data Source'!AI{data_row}>=4,\"A\",IF('Data Source'!AI{data_row}>=3.5,\"A-\",IF('Data Source'!AI{data_row}>=3,\"B\",IF('Data Source'!AI{data_row}>=2,\"C\",IF('Data Source'!AI{data_row}>=1,\"D\",\"F\"))))))"
        
        # Mathematics - use helper column AJ
        ws[f'H{row}'] = f"=IF('Data Source'!AJ{data_row}>=5,\"A+\",IF('Data Source'!AJ{data_row}>=4,\"A\",IF('Data Source'!AJ{data_row}>=3.5,\"A-\",IF('Data Source'!AJ{data_row}>=3,\"B\",IF('Data Source'!AJ{data_row}>=2,\"C\",IF('Data Source'!AJ{data_row}>=1,\"D\",\"F\"))))))"
        
        # Islamic History - use helper column AK
        ws[f'I{row}'] = f"=IF('Data Source'!AK{data_row}>=5,\"A+\",IF('Data Source'!AK{data_row}>=4,\"A\",IF('Data Source'!AK{data_row}>=3.5,\"A-\",IF('Data Source'!AK{data_row}>=3,\"B\",IF('Data Source'!AK{data_row}>=2,\"C\",IF('Data Source'!AK{data_row}>=1,\"D\",\"F\"))))))"
        
        # ICT - use helper column AL
        ws[f'J{row}'] = f"=IF('Data Source'!AL{data_row}>=5,\"A+\",IF('Data Source'!AL{data_row}>=4,\"A\",IF('Data Source'!AL{data_row}>=3.5,\"A-\",IF('Data Source'!AL{data_row}>=3,\"B\",IF('Data Source'!AL{data_row}>=2,\"C\",IF('Data Source'!AL{data_row}>=1,\"D\",\"F\"))))))"
        
        # Mantiq - use helper column AM
        ws[f'K{row}'] = f"=IF('Data Source'!AM{data_row}>=5,\"A+\",IF('Data Source'!AM{data_row}>=4,\"A\",IF('Data Source'!AM{data_row}>=3.5,\"A-\",IF('Data Source'!AM{data_row}>=3,\"B\",IF('Data Source'!AM{data_row}>=2,\"C\",IF('Data Source'!AM{data_row}>=1,\"D\",\"F\"))))))"
        
        # Career Education (Pass/Fail - T)
        ws[f'L{row}'] = f"=IF('Data Source'!T{data_row}>=33,\"Pass\",\"Fail\")"
        
        # Physical Education (Pass/Fail - U)
        ws[f'M{row}'] = f"=IF('Data Source'!U{data_row}>=33,\"Pass\",\"Fail\")"
        
        # Overall GPA from Data Source (column Y - Final GPA)
        ws[f'N{row}'] = f"='Data Source'!Y{data_row}"
        ws[f'N{row}'].number_format = '0.00'
        
        # Overall Grade from Data Source (column Z)
        ws[f'O{row}'] = f"='Data Source'!Z{data_row}"
        
        # Alignment
        for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 10
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 12

def create_subjectwise_gpa_sheet(wb, df):
    """Create Subject-wise GPA sheet displaying individual GP values for all subjects"""
    
    ws = wb.create_sheet('Subject-wise GPA')
    
    # Title
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = 'SUBJECT-WISE GPA (GRADE POINTS)'
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Headers
    headers = ['SL', 'Name', 'Quran+Hadith GP', 'Arabic GP', 'Aqaid GP', 'English GP', 
               'Bangla GP', 'Math GP', 'History GP', 'ICT GP', 'Mantiq GP', 
               'Base GPA', 'Final GPA', 'Grade']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data - reference helper columns from Data Source
    for row_idx in range(len(df)):
        row = row_idx + 3  # Start from row 3 (after header)
        data_row = row_idx + 2  # Corresponding row in Data Source
        
        # SL and Name from Data Source
        ws[f'A{row}'] = f"='Data Source'!A{data_row}"
        ws[f'B{row}'] = f"='Data Source'!B{data_row}"
        
        # Individual subject GPs from helper columns (AE-AM)
        ws[f'C{row}'] = f"='Data Source'!AE{data_row}"  # Quran+Hadith GP
        ws[f'C{row}'].number_format = '0.00'
        
        ws[f'D{row}'] = f"='Data Source'!AF{data_row}"  # Arabic GP
        ws[f'D{row}'].number_format = '0.00'
        
        ws[f'E{row}'] = f"='Data Source'!AG{data_row}"  # Aqaid GP
        ws[f'E{row}'].number_format = '0.00'
        
        ws[f'F{row}'] = f"='Data Source'!AH{data_row}"  # English GP
        ws[f'F{row}'].number_format = '0.00'
        
        ws[f'G{row}'] = f"='Data Source'!AI{data_row}"  # Bangla GP
        ws[f'G{row}'].number_format = '0.00'
        
        ws[f'H{row}'] = f"='Data Source'!AJ{data_row}"  # Math GP
        ws[f'H{row}'].number_format = '0.00'
        
        ws[f'I{row}'] = f"='Data Source'!AK{data_row}"  # Islamic History GP
        ws[f'I{row}'].number_format = '0.00'
        
        ws[f'J{row}'] = f"='Data Source'!AL{data_row}"  # ICT GP
        ws[f'J{row}'].number_format = '0.00'
        
        ws[f'K{row}'] = f"='Data Source'!AM{data_row}"  # Mantiq GP
        ws[f'K{row}'].number_format = '0.00'
        
        ws[f'L{row}'] = f"='Data Source'!AN{data_row}"  # Base GPA (compulsory only)
        ws[f'L{row}'].number_format = '0.00'
        
        ws[f'M{row}'] = f"='Data Source'!Y{data_row}"   # Final GPA (with Mantiq bonus)
        ws[f'M{row}'].number_format = '0.00'
        
        ws[f'N{row}'] = f"='Data Source'!Z{data_row}"   # Overall Grade
        
        # Center alignment for all data cells
        for col in ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col].width = 13
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 10

def create_dashboard_sheet(wb, df):
    """Create Dashboard sheet with charts and summary"""
    
    ws = wb.create_sheet('Dashboard')
    
    # Title
    ws.merge_cells('A1:L2')
    title_cell = ws['A1']
    title_cell.value = '📊 ACADEMIC RESULTS DASHBOARD'
    title_cell.font = Font(size=24, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Summary Statistics (use Excel formulas so dashboard updates when Data Source changes)
    ws['A4'] = 'SUMMARY STATISTICS'
    ws['A4'].font = Font(size=14, bold=True)

    # Total Students (count names), Average GPA, Highest Total, Pass Rate
    ws['B5'] = 'Total Students:'
    ws['C5'] = "=COUNTA('Data Source'!B:B)-1"
    ws['E5'] = 'Average GPA:'
    ws['F5'] = "=IFERROR(ROUND(AVERAGE('Data Source'!Y:Y),2), 0)"  # Column Y is Final GPA
    ws['H5'] = 'Highest Total:'
    ws['I5'] = "=MAX('Data Source'!V:V)"  # Column V is Total
    ws['K5'] = 'Pass Rate:'
    # Count GPA > 0 divided by total - format as percentage
    ws['L5'] = '=IF(C5>0,COUNTIF(\'Data Source\'!Y:Y,">0")/C5,0)'  # Column Y is Final GPA
    ws['L5'].number_format = '0.0%'

    # Style the label cells
    for label_cell in ['B5','E5','H5','K5']:
        ws[label_cell].font = Font(bold=True)
    for value_cell in ['C5','F5','I5','L5']:
        ws[value_cell].font = Font(size=12, bold=True, color="1F4E78")
    
    # Format number cells
    ws['F5'].number_format = '0.00'  # GPA with 2 decimals
    
    # Grade Distribution Table
    ws['A7'] = 'GRADE DISTRIBUTION'
    ws['A7'].font = Font(size=12, bold=True)
    
    # Grade distribution using COUNTIF for common grade buckets
    ws['A8'] = 'Grade'
    ws['B8'] = 'Count'
    ws['A8'].font = Font(bold=True)
    ws['B8'].font = Font(bold=True)
    ws['A8'].alignment = Alignment(horizontal='center')
    ws['B8'].alignment = Alignment(horizontal='center')
    
    grade_list = ['A+','A','A-','B','C','D','F']
    row = 9
    for grade in grade_list:
        ws[f'A{row}'] = grade
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        # COUNTIF on Data Source Overall Grade column Z
        ws[f'B{row}'] = f"=COUNTIF('Data Source'!Z:Z,\"{grade}\")"
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # Subject-wise Average
    ws['D7'] = 'SUBJECT-WISE AVERAGE'
    ws['D7'].font = Font(size=12, bold=True)
    
    # Dakhil curriculum subjects with their columns and full marks
    subjects_info = [
        ('Quran Mazid', 'C', 100, False),
        ('Arabic (Comb.)', ['E', 'F'], 200, True),
        ('Aqaid', 'G', 100, False),
        ('English (Comb.)', ['H', 'I'], 200, True),
        ('Bangla (Comb.)', ['J', 'K', 'L', 'M'], 200, True),
        ('Mathematics', ['N', 'O'], 100, True),
        ('Islamic History', ['P', 'Q'], 100, True),
        ('ICT', 'R', 50, False),
        ('Mantiq', 'S', 100, False)
    ]
    
    ws['D8'] = 'Subject'
    ws['E8'] = 'Avg'
    ws['F8'] = '%'
    ws['D8'].font = Font(bold=True)
    ws['E8'].font = Font(bold=True)
    ws['F8'].font = Font(bold=True)
    ws['E8'].alignment = Alignment(horizontal='center')
    ws['F8'].alignment = Alignment(horizontal='center')
    
    row = 9
    for subject_name, col_info, full_marks, is_multi_col in subjects_info:
        ws[f'D{row}'] = subject_name
        
        if is_multi_col:
            # For multi-column subjects, we need to sum columns first
            cols = '+'.join([f"'Data Source'!{col}3:{col}22" for col in col_info])
            ws[f'E{row}'] = f"=IFERROR(ROUND(AVERAGE({cols}),2),0)"
        else:
            # Single column subjects
            ws[f'E{row}'] = f"=IFERROR(ROUND(AVERAGE('Data Source'!{col_info}:{col_info}),2),0)"
        
        ws[f'E{row}'].number_format = '0.00'
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        # Percentage average
        ws[f'F{row}'] = f"=IFERROR(ROUND(E{row}/{full_marks}*100,1),0)"
        ws[f'F{row}'].number_format = '0.0"%"'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # Top 5 Students - Simplified approach using Python to get initial top 5
    # Then use formulas that reference Data Source directly
    ws['G7'] = 'TOP 5 STUDENTS'
    ws['G7'].font = Font(size=12, bold=True)
    
    ws['G8'] = 'Rank'
    ws['H8'] = 'Name'
    ws['I8'] = 'GPA'
    ws['G8'].font = Font(bold=True)
    ws['H8'].font = Font(bold=True)
    ws['I8'].font = Font(bold=True)
    ws['G8'].alignment = Alignment(horizontal='center')
    ws['I8'].alignment = Alignment(horizontal='center')
    
    # Calculate top 5 students in Python to get their row numbers using Dakhil GPA calculation
    df_with_gpa = df.copy()
    df_with_gpa['GPA_calc'] = df_with_gpa.apply(calculate_gpa_dakhil, axis=1)
    df_sorted = df_with_gpa.nlargest(5, 'GPA_calc')
    top_5_rows = [df_with_gpa.index.get_loc(idx) + 2 for idx in df_sorted.index]  # +2 for header and 1-based indexing
    
    # Create Top 5 list with formulas that reference specific rows from Data Source
    row = 9
    for i, data_source_row in enumerate(top_5_rows, 1):
        ws[f'G{row}'] = i
        ws[f'G{row}'].alignment = Alignment(horizontal='center')
        
        # Reference the specific student's name and GPA from Data Source (column Y = Final GPA)
        # This way if their marks change, their GPA updates
        ws[f'H{row}'] = f"='Data Source'!B{data_source_row}"
        ws[f'I{row}'] = f"='Data Source'!Y{data_source_row}"
        ws[f'I{row}'].number_format = '0.00'
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Create Pie Chart - Grade Distribution
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=9, max_row=15)
    data = Reference(ws, min_col=2, min_row=8, max_row=15)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Grade Distribution"
    ws.add_chart(pie, "A17")
    
    # Create Bar Chart - Subject-wise Performance (9 Dakhil subjects)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Subject-wise Average Scores"
    bar.y_axis.title = 'Average Marks'
    bar.x_axis.title = 'Subjects'
    
    data = Reference(ws, min_col=5, min_row=8, max_row=17)  # E8:E17 (9 subjects + header)
    cats = Reference(ws, min_col=4, min_row=9, max_row=17)  # D9:D17 (9 subject names)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws.add_chart(bar, "G17")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['H'].width = 18

def generate_excel_file(filename='Academic_Results_Dashboard.xlsx'):
    """Main function to generate the Excel file"""
    
    print("🚀 Generating Academic Results Dashboard...")
    
    # Create data
    df = create_data_source()
    
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Source'
    
    # Write data to Data Source sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Add Total, Average, GPA, Overall Grade columns (columns V-Y)
    # Subject columns: C-U (19 subjects - includes split MCQ+Written for Bangla, Math, Islamic History)
    # Summary columns: V=Total, W=Average, X=GPA (Compulsory), Y=GPA (Final), Z=Overall Grade
    
    # Summary columns
    ws['V1'] = 'Total'
    ws['W1'] = 'Average'
    ws['X1'] = 'GPA (Compulsory)'
    ws['Y1'] = 'GPA (Final)'
    ws['Z1'] = 'Overall Grade'
    
    # Add formulas for each student row (rows 2 to len(df)+1)
    for row in range(2, len(df) + 2):
        # New column structure:
        # C=Quran(100), D=Hadith(100), E=Arabic I(100), F=Arabic II(100), G=Aqaid(100), 
        # H=English I(100), I=English II(100), J=Bangla_I_MCQ(30), K=Bangla_I_Written(70), 
        # L=Bangla_II_MCQ(30), M=Bangla_II_Written(70), N=Math_MCQ(30), O=Math_Written(70),
        # P=Islamic_History_MCQ(30), Q=Islamic_History_Written(70), R=ICT(50), 
        # S=Mantiq(100), T=Career(100), U=Physical(100)
        
        # Total: Sum of compulsory subjects (8 subjects but 16 columns due to splits)
        # Compulsory: Quran+Hadith(200), Arabic I+II(200), Aqaid(100), English I+II(200), 
        # Bangla (4 cols, 200), Math (2 cols, 100), Islamic History (2 cols, 100), ICT(50)
        # That's columns C through R (excluding S=Mantiq, T=Career, U=Physical)
        # Total marks = 1150
        ws[f'V{row}'] = f'=SUM(C{row}:R{row})'
        
        # Average: Total/11.5 to normalize (1150/11.5 = 100-mark equivalent)
        ws[f'W{row}'] = f'=ROUND(V{row}/11.5,2)'
        ws[f'W{row}'].number_format = '0.00'
        
        # SIMPLIFIED APPROACH: Use helper columns to avoid formula corruption
        # Instead of one massive formula, break it into manageable pieces
        
        # Helper columns for fail conditions (columns AA-AD, hidden later)
        # AA: Bangla fail check
        ws[f'AA{row}'] = f'=NOT(OR(AND(J{row}>=10,K{row}>=23,L{row}>=10,M{row}>=23),AND(J{row}+L{row}>=20,K{row}+M{row}>=46)))'
        # AB: Math fail check
        ws[f'AB{row}'] = f'=OR(N{row}<10,O{row}<23)'
        # AC: History fail check
        ws[f'AC{row}'] = f'=OR(P{row}<10,Q{row}<23)'
        # AD: Overall fail check (any subject failed)
        ws[f'AD{row}'] = f'=OR((C{row}+D{row})<66,(E{row}+F{row})<66,G{row}<33,(H{row}+I{row})<66,AA{row},AB{row},AC{row},R{row}<8.25,T{row}<33,U{row}<33)'
        
        # Helper columns for grade points - simpler formulas
        # AE: Quran+Hadith GP (check pass threshold first)
        pct = f'(C{row}+D{row})/2'
        ws[f'AE{row}'] = f'=IF(C{row}+D{row}<66,0,IF({pct}>=80,5,IF({pct}>=70,4,IF({pct}>=60,3.5,IF({pct}>=50,3,IF({pct}>=40,2,IF({pct}>=33,1,0)))))))'
        # AF: Arabic GP (check pass threshold first)
        pct = f'(E{row}+F{row})/2'
        ws[f'AF{row}'] = f'=IF(E{row}+F{row}<66,0,IF({pct}>=80,5,IF({pct}>=70,4,IF({pct}>=60,3.5,IF({pct}>=50,3,IF({pct}>=40,2,IF({pct}>=33,1,0)))))))'
        # AG: Aqaid GP (check pass threshold first)
        ws[f'AG{row}'] = f'=IF(G{row}<33,0,IF(G{row}>=80,5,IF(G{row}>=70,4,IF(G{row}>=60,3.5,IF(G{row}>=50,3,IF(G{row}>=40,2,IF(G{row}>=33,1,0)))))))'
        # AH: English GP (check pass threshold first)
        pct = f'(H{row}+I{row})/2'
        ws[f'AH{row}'] = f'=IF(H{row}+I{row}<66,0,IF({pct}>=80,5,IF({pct}>=70,4,IF({pct}>=60,3.5,IF({pct}>=50,3,IF({pct}>=40,2,IF({pct}>=33,1,0)))))))'
        # AI: Bangla GP (check fail condition first)
        pct = f'(J{row}+K{row}+L{row}+M{row})/2'
        ws[f'AI{row}'] = f'=IF(AA{row},0,IF({pct}>=80,5,IF({pct}>=70,4,IF({pct}>=60,3.5,IF({pct}>=50,3,IF({pct}>=40,2,IF({pct}>=33,1,0)))))))'
        # AJ: Math GP (check fail condition first)
        ws[f'AJ{row}'] = f'=IF(AB{row},0,IF(N{row}+O{row}>=80,5,IF(N{row}+O{row}>=70,4,IF(N{row}+O{row}>=60,3.5,IF(N{row}+O{row}>=50,3,IF(N{row}+O{row}>=40,2,IF(N{row}+O{row}>=33,1,0)))))))'
        # AK: History GP (check fail condition first)
        ws[f'AK{row}'] = f'=IF(AC{row},0,IF(P{row}+Q{row}>=80,5,IF(P{row}+Q{row}>=70,4,IF(P{row}+Q{row}>=60,3.5,IF(P{row}+Q{row}>=50,3,IF(P{row}+Q{row}>=40,2,IF(P{row}+Q{row}>=33,1,0)))))))'
        # AL: ICT GP (check pass threshold first)
        pct = f'R{row}*2'  # Convert 50 to 100 scale
        ws[f'AL{row}'] = f'=IF(R{row}<8.25,0,IF({pct}>=80,5,IF({pct}>=70,4,IF({pct}>=60,3.5,IF({pct}>=50,3,IF({pct}>=40,2,IF({pct}>=33,1,0)))))))'
        # AM: Mantiq GP (check pass threshold first)
        ws[f'AM{row}'] = f'=IF(S{row}<33,0,IF(S{row}>=80,5,IF(S{row}>=70,4,IF(S{row}>=60,3.5,IF(S{row}>=50,3,IF(S{row}>=40,2,IF(S{row}>=33,1,0)))))))'
        
        # AN: Base GPA (average of 8 compulsory subjects)
        ws[f'AN{row}'] = f'=(AE{row}+AF{row}+AG{row}+AH{row}+AI{row}+AJ{row}+AK{row}+AL{row})/8'
        
        # AO: Mantiq bonus
        ws[f'AO{row}'] = f'=IF(AM{row}>=2,(AM{row}-2)/8,0)'
        
        # X: GPA without optional (just compulsory subjects average)
        ws[f'X{row}'] = f'=IF(AD{row},0,ROUND(AN{row},2))'
        ws[f'X{row}'].number_format = '0.00'
        
        # Y: Final GPA (with optional subject bonus)
        ws[f'Y{row}'] = f'=IF(AD{row},0,MIN(5,ROUND(AN{row}+AO{row},2)))'
        ws[f'Y{row}'].number_format = '0.00'
        
        # Z: Overall Grade (based on final GPA with optional)
        ws[f'Z{row}'] = f'=IF(Y{row}>=5,"A+",IF(Y{row}>=4,"A",IF(Y{row}>=3.5,"A-",IF(Y{row}>=3,"B",IF(Y{row}>=2,"C",IF(Y{row}>=1,"D","F"))))))'
        
        # Center align GPA and Overall Grade
        ws[f'X{row}'].alignment = Alignment(horizontal='center')
        ws[f'Y{row}'].alignment = Alignment(horizontal='center')
        ws[f'Z{row}'].alignment = Alignment(horizontal='center')
    
    # Style Data Source sheet
    style_data_source_sheet(ws, df)
    
    # Create Dashboard sheet
    create_dashboard_sheet(wb, df)
    
    # Create Subject Grades sheet
    create_subject_grades_sheet(wb, df)
    
    # Create Subject-wise GPA sheet
    create_subjectwise_gpa_sheet(wb, df)
    
    # Note: Pivot sheet would require manual creation in Excel or additional library
    wb.create_sheet('Pivot')
    pivot_ws = wb['Pivot']
    pivot_ws['A1'] = 'Pivot tables can be created manually in Excel'
    pivot_ws['A2'] = 'Use Insert > PivotTable from the Data Source sheet'
    
    # Set calculation mode to automatic to ensure formulas calculate on open
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
    
    # Save file
    wb.save(filename)
    print(f"✅ Excel file created successfully: {filename}")
    # For console summary we can compute GPA locally (this does not alter the
    # workbook which contains formulas). This provides quick feedback to user.
    df_console = df.copy()
    try:
        df_console['GPA'] = df_console.apply(calculate_gpa_dakhil, axis=1)
        # Grade based on GPA
        def get_grade(gpa):
            if gpa >= 5: return 'A+'
            elif gpa >= 4: return 'A'
            elif gpa >= 3.5: return 'A-'
            elif gpa >= 3: return 'B'
            elif gpa >= 2: return 'C'
            elif gpa >= 1: return 'D'
            else: return 'F'
        df_console['Grade'] = df_console['GPA'].apply(get_grade)
        avg_gpa = df_console['GPA'].mean()
        num_a_plus = (df_console['Grade'] == 'A+').sum()
        pass_rate = (df_console['GPA'] > 0).sum() / len(df_console) * 100
    except Exception:
        avg_gpa = 0
        num_a_plus = 0
        pass_rate = 0

    print(f"\n📊 Summary:")
    print(f"   - Total Students: {len(df)}")
    print(f"   - Average Class GPA: {avg_gpa:.2f}")
    print(f"   - Students with A+: {num_a_plus}")
    print(f"   - Pass Rate: {pass_rate:.1f}%")
    print(f"\n📁 File saved as: {filename}")

if __name__ == "__main__":
    generate_excel_file()