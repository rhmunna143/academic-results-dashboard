import openpyxl

# Load without data_only to see formulas
wb = openpyxl.load_workbook('Academic_Results_Dashboard.xlsx')
ws = wb['Data Source']

print('Checking formula in X3 (GPA):')
print(ws['X3'].value)
print()
print('Checking formula in Y3 (Grade):')
print(ws['Y3'].value)
print()
print('Checking some data values:')
print(f'C3 (Quran): {ws["C3"].value}')
print(f'D3 (Hadith): {ws["D3"].value}')
print(f'J3 (Bangla_I_MCQ): {ws["J3"].value}')
print(f'K3 (Bangla_I_Written): {ws["K3"].value}')
print(f'L3 (Bangla_II_MCQ): {ws["L3"].value}')
print(f'M3 (Bangla_II_Written): {ws["M3"].value}')
print(f'N3 (Math_MCQ): {ws["N3"].value}')
print(f'O3 (Math_Written): {ws["O3"].value}')
print(f'R3 (ICT): {ws["R3"].value}')

wb.close()
